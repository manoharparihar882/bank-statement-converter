#!/usr/bin/env python3
"""
PDF to Excel converter for Indian bank statements
"""

import sys
import os
import json
import pandas as pd
import tabula
import pdfplumber
import re
import camelot
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def extract_sbi_data(pdf_path):
    """Extract data from SBI bank statement"""
    tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
    
    # Look for transaction tables
    transactions = []
    for table in tables:
        # SBI specific validation - checking column names
        if 'Date' in table.columns and 'Description' in table.columns:
            transactions.append(table)
    
    # Combine all transaction tables
    if transactions:
        df = pd.concat(transactions, ignore_index=True)
        # Clean up columns, remove any unnamed columns
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        return df
    
    return pd.DataFrame()

def extract_hdfc_data(pdf_path):
    """Extract data from HDFC bank statement"""
    # HDFC often requires custom parsing with pdfplumber
    transactions = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Extract tables from the page
            tables = page.extract_tables()
            
            for table in tables:
                # Skip empty tables
                if not table or len(table) <= 1:
                    continue
                
                # Check if this looks like a transaction table
                header = table[0]
                if any(col and 'Date' in str(col) for col in header):
                    # Create DataFrame from the table, skipping header
                    df = pd.DataFrame(table[1:], columns=table[0])
                    transactions.append(df)
    
    if transactions:
        df = pd.concat(transactions, ignore_index=True)
        return df
    
    return pd.DataFrame()

def extract_icici_data(pdf_path):
    """Extract data from ICICI bank statement"""
    # ICICI often has better results with Camelot
    tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
    
    transactions = []
    for table in tables:
        df = table.df
        # ICICI specific validation - checking if it looks like a transaction table
        if len(df.columns) >= 5 and any('Date' in str(col) for col in df.iloc[0]):
            # Use the first row as header
            df.columns = df.iloc[0]
            df = df.iloc[1:]
            transactions.append(df)
    
    if transactions:
        df = pd.concat(transactions, ignore_index=True)
        return df
    
    return pd.DataFrame()

def detect_bank_type(pdf_path):
    """Detect bank type from statement"""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages[:2]:  # Check first two pages
            text += page.extract_text().lower()
    
    if "state bank of india" in text or "sbi" in text:
        return "SBI"
    elif "hdfc bank" in text:
        return "HDFC"
    elif "icici bank" in text:
        return "ICICI"
    elif "axis bank" in text:
        return "AXIS"
    elif "punjab national bank" in text or "pnb" in text:
        return "PNB"
    elif "kotak mahindra bank" in text:
        return "KOTAK"
    elif "bank of baroda" in text:
        return "BOB"
    elif "union bank" in text:
        return "UNION"
    elif "canara bank" in text:
        return "CANARA"
    else:
        return "GENERIC"

def cleanup_and_format_data(df):
    """Cleanup and standardize dataframe format"""
    # Standardize column names
    column_map = {
        'Txn Date': 'Date',
        'Transaction Date': 'Date',
        'Value Date': 'Date',
        'Particulars': 'Description',
        'Details': 'Description',
        'Transaction Details': 'Description',
        'Narration': 'Description',
        'Withdrawal Amt.': 'Debit',
        'Withdrawals': 'Debit',
        'Debit Amount': 'Debit',
        'Amount (Rs.)': 'Debit',
        'Deposit Amt.': 'Credit',
        'Deposits': 'Credit',
        'Credit Amount': 'Credit',
        'Balance (Rs.)': 'Balance',
    }
    
    # Rename columns based on the map, ignoring any not in the map
    df = df.rename(columns={col: column_map.get(col, col) for col in df.columns})
    
    # Ensure these columns exist
    required_columns = ['Date', 'Description', 'Debit', 'Credit', 'Balance']
    for col in required_columns:
        if col not in df.columns:
            df[col] = ""
    
    # Convert numeric columns
    for col in ['Debit', 'Credit', 'Balance']:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).str.replace(',', '').str.replace('₹', '').str.replace('Rs.', '').str.replace('-', '0')
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    # Format date
    if df['Date'].dtype == object:
        df['Date'] = pd.to_datetime(df['Date'], format='mixed', errors='coerce')
        df['Date'] = df['Date'].dt.strftime('%d-%m-%Y')
    
    # Drop rows with NaN date (usually headers or footers)
    df = df.dropna(subset=['Date'])
    
    # Ensure "Cheque No." column exists
    if 'Cheque No.' not in df.columns:
        df['Cheque No.'] = ""
    
    # Reorder columns
    ordered_columns = ['Date', 'Description', 'Cheque No.', 'Debit', 'Credit', 'Balance']
    return df[ordered_columns]

def write_to_excel(df, output_path):
    """Write dataframe to Excel with formatting"""
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"
    
    # Add headers
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx+2, column=col_idx, value=value)
            # Format numbers
            if col_idx in [4, 5, 6]:  # Debit, Credit, Balance columns
                cell.number_format = '₹#,##0.00'
    
    # Auto-adjust column width
    for col_idx, column in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_idx)
        # Calculate the maximum width needed
        max_length = max(
            len(str(cell.value)) for cell in ws[column_letter]
        )
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_path)
    
    return True

def main():
    if len(sys.argv) != 3:
        print("Usage: python pdf_to_excel.py <pdf_path> <output_path>")
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    output_path = sys.argv[2]
    
    # Check if input file exists
    if not os.path.exists(pdf_path):
        print(f"Error: File {pdf_path} not found")
        sys.exit(1)
    
    try:
        # Detect bank type
        bank_type = detect_bank_type(pdf_path)
        print(f"Detected bank type: {bank_type}")
        
        # Extract data based on bank type
        if bank_type == "SBI":
            df = extract_sbi_data(pdf_path)
        elif bank_type == "HDFC":
            df = extract_hdfc_data(pdf_path)
        elif bank_type == "ICICI":
            df = extract_icici_data(pdf_path)
        else:
            # Try different methods and use the one with most data
            methods = [
                ("SBI method", extract_sbi_data(pdf_path)),
                ("HDFC method", extract_hdfc_data(pdf_path)),
                ("ICICI method", extract_icici_data(pdf_path))
            ]
            
            # Find method with most rows
            df = max(methods, key=lambda x: len(x[1]))[1]
        
        # Clean up and format data
        df = cleanup_and_format_data(df)
        
        # Write to Excel
        write_to_excel(df, output_path)
        
        # Output preview data as JSON for Node.js to parse
        preview_data = df.head(10).to_dict('records')
        print(json.dumps(preview_data))
        
        sys.exit(0)
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main() 